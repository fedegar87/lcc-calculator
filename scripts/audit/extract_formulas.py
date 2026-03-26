"""Extract all formula cells from all 7 sheets of the CRAVEzero workbook using dual-pass technique."""

import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

WORKBOOK = Path("CRAVEzero/200512_LCC_tool_beta_v2.xlsm")
OUTPUT = Path("scripts/output/formulas_raw.json")

SHEETS = [
    "Project Information",
    "WLC",
    "Construction cost",
    "Maintenance",
    "Results",
    "Calc",
    "Charts",
]


class FormulaEncoder(json.JSONEncoder):
    """Handle datetime and other non-serializable cached values."""

    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


def extract():
    # Pass 1: formula text
    wb_formulas = openpyxl.load_workbook(WORKBOOK, data_only=False, keep_vba=True)
    # Pass 2: cached computed values
    wb_values = openpyxl.load_workbook(WORKBOOK, data_only=True, keep_vba=True)

    sheets_data = {}
    total_formulas = 0

    for sheet_name in SHEETS:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]

        formulas = []
        for row in range(1, ws_f.max_row + 1):
            for col in range(1, ws_f.max_column + 1):
                cell_f = ws_f.cell(row=row, column=col)
                if cell_f.value and isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    cell_v = ws_v.cell(row=row, column=col)
                    formulas.append({
                        "sheet": sheet_name,
                        "cell": cell_f.coordinate,
                        "row": row,
                        "col": col,
                        "formula": cell_f.value,
                        "cached_value": cell_v.value,
                    })

        sheets_data[sheet_name] = {
            "count": len(formulas),
            "formulas": formulas,
        }
        total_formulas += len(formulas)
        print(f"  {sheet_name}: {len(formulas)} formula cells")

    # Assertions
    assert len(sheets_data) == 7, f"Expected 7 sheets, got {len(sheets_data)}"
    assert total_formulas > 500, f"Expected >500 formulas, got {total_formulas}"
    assert sheets_data["Calc"]["count"] > 200, (
        f"Calc should have >200 formulas, got {sheets_data['Calc']['count']}"
    )
    assert sheets_data["Results"]["count"] > 100, (
        f"Results should have >100 formulas, got {sheets_data['Results']['count']}"
    )

    envelope = {
        "source": "CRAVEzero workbook formula extraction",
        "workbook": str(WORKBOOK),
        "extraction_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "sheets": sheets_data,
        "total_formulas": total_formulas,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(envelope, f, indent=2, ensure_ascii=False, cls=FormulaEncoder)

    print(f"\nTotal: {total_formulas} formulas across {len(sheets_data)} sheets")
    print(f"Written to {OUTPUT}")


if __name__ == "__main__":
    extract()
