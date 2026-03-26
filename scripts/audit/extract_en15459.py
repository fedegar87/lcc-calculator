"""Extract EN 15459 HVAC component table from CRAVEzero workbook (Calc!B405:H483)."""

import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

WORKBOOK = Path("CRAVEzero/200512_LCC_tool_beta_v2.xlsm")
OUTPUT = Path("scripts/output/en15459.json")


def extract():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, keep_vba=True)
    ws = wb["Calc"]

    components = []
    for row in range(405, 484):
        name = ws.cell(row=row, column=2).value  # B
        if name is None:
            continue
        components.append({
            "index": row - 404,
            "name": name,
            "lifespan_min": ws.cell(row=row, column=3).value,
            "lifespan_max": ws.cell(row=row, column=4).value,
            "lifespan_avg": ws.cell(row=row, column=5).value,
            "maintenance_pct_min": ws.cell(row=row, column=6).value,
            "maintenance_pct_max": ws.cell(row=row, column=7).value,
            "maintenance_pct_avg": ws.cell(row=row, column=8).value,
        })

    # Assertions
    assert len(components) == 79, f"Expected 79 EN 15459 components, got {len(components)}"
    assert components[0]["name"] == "Air conditioning units", f"First component: {components[0]['name']}"
    assert components[-1]["name"] == "Wiring", f"Last component: {components[-1]['name']}"
    assert all(c["lifespan_avg"] > 0 for c in components), "All lifespans must be positive"
    for c in components:
        if c["maintenance_pct_avg"] is not None:
            assert 0 <= c["maintenance_pct_avg"] <= 0.15, (
                f"{c['name']}: maintenance_pct_avg={c['maintenance_pct_avg']} out of range 0-0.15"
            )

    envelope = {
        "source": "EN 15459:2018 via CRAVEzero workbook",
        "extracted_from": "Calc!B405:H483",
        "extraction_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "count": len(components),
        "components": components,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(envelope, f, indent=2, ensure_ascii=False)

    print(f"Extracted {len(components)} EN 15459 components to {OUTPUT}")


if __name__ == "__main__":
    extract()
